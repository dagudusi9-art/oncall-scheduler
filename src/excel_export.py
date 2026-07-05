# -*- coding: utf-8 -*-
"""
最適化結果・実績・年間集計・交代履歴をExcelファイルへ出力する。

出力シート(与えられたデータに応じて作成される。すべて省略可能):
  「予定勤務表」   : 予定(scheduled_assignments)の日付ごとの担当者一覧
  「実績勤務表」   : 実績(actual_assignments)の日付ごとの担当者一覧
  「月間集計」     : 実績があれば実績ベース、無ければ予定ベースの集計
  「年間実績集計」 : 年間の自院日中/夜間/合計・外部バイト・総勤務の実績集計
  「交代履歴」     : 勤務交代の申請・承認履歴
  「警告」         : 割当時の警告・注意事項(あれば)
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Union

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import ScheduleEntry, ScheduleResult, Slot

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WEEKEND_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

WEEKDAY_JA = ["月", "火", "水", "木", "金", "土", "日"]
SLOT_TYPE_LABEL = {"day": "日中", "night": "夜間", "gaikobu": "外部バイト"}
SWAP_STATUS_LABEL = {"pending": "保留中", "approved": "承認済み", "rejected": "却下"}


def _normalize_entries(entries: List[Union[ScheduleEntry, dict]]) -> List[dict]:
    """ScheduleEntryのリスト、またはdict(スナップショット由来)のリストを
    共通の {"date": date, "day":.., "night":.., "gaikobu":..} 形式に揃える。"""
    from datetime import datetime

    normalized = []
    for e in entries:
        if isinstance(e, ScheduleEntry):
            normalized.append(
                {
                    "date": e.day,
                    "day": e.assignments.get(Slot.DAY),
                    "night": e.assignments.get(Slot.NIGHT),
                    "gaikobu": e.gaikobu,
                }
            )
        else:
            d = e["date"]
            if isinstance(d, str):
                d = datetime.strptime(d, "%Y-%m-%d").date()
            normalized.append({"date": d, "day": e.get("day"), "night": e.get("night"), "gaikobu": e.get("gaikobu")})
    return normalized


def export_to_excel(
    scheduled_result: ScheduleResult,
    output_path: str | Path,
    actual_entries: Optional[List[dict]] = None,
    actual_stats: Optional[Dict[str, Dict[str, int]]] = None,
    annual_totals: Optional[Dict[str, Dict[str, int]]] = None,
    swap_history: Optional[List[dict]] = None,
) -> Path:
    """
    scheduled_result: 予定(必須。既存のScheduleResult)
    actual_entries / actual_stats: 実績(あれば「実績勤務表」シートを追加し、
        「月間集計」は実績ベースになる。無ければ予定ベースで集計する)
    annual_totals: 年間実績集計(あれば「年間実績集計」シートを追加)
    swap_history: 交代履歴のリスト(あれば「交代履歴」シートを追加)
    """
    output_path = Path(output_path)
    wb = Workbook()

    _write_calendar_sheet(wb, "予定勤務表", scheduled_result.entries)
    if actual_entries is not None:
        _write_calendar_sheet(wb, "実績勤務表", actual_entries)

    monthly_stats = actual_stats if actual_stats is not None else scheduled_result.stats
    _write_stats_sheet(wb, "月間集計", monthly_stats)

    if annual_totals:
        _write_annual_sheet(wb, annual_totals)

    if swap_history:
        _write_swap_history_sheet(wb, swap_history)

    if scheduled_result.warnings:
        _write_warnings_sheet(wb, scheduled_result.warnings)

    # デフォルトで作成される空シートを削除
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _write_calendar_sheet(wb: Workbook, sheet_name: str, entries: List[Union[ScheduleEntry, dict]]) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = ["日付", "曜日", "日中", "夜間", "外部バイト"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row, e in enumerate(_normalize_entries(entries), start=2):
        day = e["date"]
        weekday_idx = day.weekday()
        ws.cell(row=row, column=1, value=f"{day.month}/{day.day}")
        ws.cell(row=row, column=2, value=WEEKDAY_JA[weekday_idx])
        day_name = e.get("day") or "(未割当)"
        night_name = e.get("night") or "(未割当)"
        gaikobu_name = e.get("gaikobu") or ""
        c_day = ws.cell(row=row, column=3, value=day_name)
        c_night = ws.cell(row=row, column=4, value=night_name)
        ws.cell(row=row, column=5, value=gaikobu_name)

        if weekday_idx >= 5:  # 土日は色付け
            for c in (ws.cell(row=row, column=1), ws.cell(row=row, column=2)):
                c.fill = WEEKEND_FILL

        if day_name == "(未割当)":
            c_day.fill = WARN_FILL
        if night_name == "(未割当)":
            c_night.fill = WARN_FILL

    for col, width in zip("ABCDE", (10, 6, 14, 14, 14)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _write_stats_sheet(wb: Workbook, sheet_name: str, stats: Dict[str, Dict[str, int]]) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = ["名前", "日中", "夜間", "自院合計", "外部バイト", "総勤務", "目標", "差"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row, (name, s) in enumerate(stats.items(), start=2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=s.get("day", 0))
        ws.cell(row=row, column=3, value=s.get("night", 0))
        ws.cell(row=row, column=4, value=s.get("total", 0))
        ws.cell(row=row, column=5, value=s.get("gaikobu", 0))
        ws.cell(row=row, column=6, value=s.get("grand_total", s.get("total", 0)))
        ws.cell(row=row, column=7, value=s.get("target", 0))
        diff_cell = ws.cell(row=row, column=8, value=s.get("diff", 0))
        if s.get("diff", 0) != 0:
            diff_cell.fill = WARN_FILL

    for col, width in zip("ABCDEFGH", (12, 8, 8, 10, 10, 8, 8, 6)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _write_annual_sheet(wb: Workbook, annual_totals: Dict[str, Dict[str, int]]) -> None:
    ws = wb.create_sheet("年間実績集計")
    headers = ["名前", "自院日中 実績", "自院夜間 実績", "自院合計 実績", "外部バイト 実績", "総勤務 実績"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row, (name, s) in enumerate(annual_totals.items(), start=2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=s.get("day", 0))
        ws.cell(row=row, column=3, value=s.get("night", 0))
        ws.cell(row=row, column=4, value=s.get("total", 0))
        ws.cell(row=row, column=5, value=s.get("gaikobu", 0))
        ws.cell(row=row, column=6, value=s.get("grand_total", 0))

    for col, width in zip("ABCDEF", (12, 14, 14, 14, 14, 12)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _write_swap_history_sheet(wb: Workbook, swap_history: List[dict]) -> None:
    ws = wb.create_sheet("交代履歴")
    headers = ["日付", "勤務種別", "元の担当者", "交代後の担当者", "申請日時", "承認日時", "状態"]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(swap_history, start=2):
        ws.cell(row=row, column=1, value=r.get("date", ""))
        ws.cell(row=row, column=2, value=SLOT_TYPE_LABEL.get(r.get("slot_type"), r.get("slot_type", "")))
        ws.cell(row=row, column=3, value=r.get("from_member", ""))
        ws.cell(row=row, column=4, value=r.get("to_member", ""))
        ws.cell(row=row, column=5, value=(r.get("requested_at") or "")[:16].replace("T", " "))
        ws.cell(
            row=row,
            column=6,
            value=(r.get("approved_at") or "")[:16].replace("T", " ") if r.get("approved_at") else "",
        )
        ws.cell(row=row, column=7, value=SWAP_STATUS_LABEL.get(r.get("status"), r.get("status", "")))

    for col, width in zip("ABCDEFG", (12, 12, 14, 14, 18, 18, 10)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"


def _write_warnings_sheet(wb: Workbook, warnings: List[str]) -> None:
    ws = wb.create_sheet("警告")
    ws.cell(row=1, column=1, value="割当時の警告・注意事項").font = Font(bold=True)
    for row, msg in enumerate(warnings, start=2):
        ws.cell(row=row, column=1, value=msg)
    ws.column_dimensions["A"].width = 80
